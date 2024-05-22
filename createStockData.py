import warnings
import yfinance as yf
from openpyxl import Workbook
import StockCharts

# Ignore FutureWarning
warnings.simplefilter(action='ignore', category=FutureWarning)

#---------------------------------------------------------------------------------------------------------
# No longer used with the main file
# Still works to add stock data in an excel file
#----------------------------------------------------------------------------------------------------------

# Creating an Excel Sheet
wb = Workbook()
ws = wb.active

# Stock name
stock = ""

# Create a Ticker object for the stock
stockData = yf.Ticker(stock)

# Get the historical market data
hist = stockData.history(period="5y")

# Remove timezone from dates
hist.index = hist.index.tz_localize(None)

# Add dates to excel sheet
ws['A1'] = "Date"
for i in range(len(hist)):
    ws['A'+ str(i+2)] = hist.index[i]

# Add closing prices to excel sheet
ws['B1'] = "Closing"
for i in range(len(hist)):
    ws['B'+ str(i+2)] = hist['Close'].iloc[i]

# Add percentage change
ws['C1'] = "Pct Change"
hist['Percentage Change'] = hist['Close'].pct_change() * 100
for i in range(len(hist)):
    ws['C'+ str(i+2)] = hist['Percentage Change'].iloc[i]


wb.save('./Data/' + stock + 'Data.xlsx')
print("New stock added: " + stock)