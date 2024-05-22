import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
import os
import numpy as np
import tkinter as tk
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import warnings
import yfinance as yf
from openpyxl import Workbook
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression
from sklearn import preprocessing
from datetime import datetime

def getStockData(stock):
    # Ignore FutureWarning
    warnings.simplefilter(action='ignore', category=FutureWarning)

    # Creating an Excel Sheet
    wb = Workbook()
    ws = wb.active

    # Create a Ticker object for the stock
    stockData = yf.Ticker(stock)

    # Get the historical market data
    hist = stockData.history(period="5y")
    
    # Remove timezone from dates
    hist.index = hist.index.tz_localize(None)

    # Add dates to excel sheet
    ws['A1'] = "Date"
    for i in range(len(hist)):
        ws['A'+ str(i+2)] = hist.index[i]

    # Add closing prices to excel sheet
    ws['B1'] = "Closing"
    for i in range(len(hist)):
        ws['B'+ str(i+2)] = hist['Close'].iloc[i]

    # Add percentage change
    ws['C1'] = "Pct Change"
    hist['Percentage Change'] = hist['Close'].pct_change() * 100
    for i in range(len(hist)):
        ws['C'+ str(i+2)] = hist['Percentage Change'].iloc[i]

    wb.save('./Data/' + stock + 'Data.xlsx')
    print("New stock added/updated: " + stock)

# ----------------------------------------------------------------
# create the window
def create_popup(stock, length):
    for widget in chartFrame.winfo_children():
        widget.destroy()

    fig = Figure(figsize=(5, 5), dpi=100)
    ax = fig.add_subplot(111)

    # Load spreadsheet
    xl = pd.ExcelFile('./Data/' + stock + 'Data.xlsx')
    df = xl.parse('Sheet')

    colors = np.where(df['Closing'].diff() > 0, 'green', 'red')
    ax.set_facecolor('darkgray')
    ax.set_title(stock)

    
    #button = tk.Button(chartFrame, text='5 yr', command=lambda: create_popup(stock, 0))
    #button.pack()
    #button = tk.Button(chartFrame, text='1 yr', command=lambda: create_popup(stock, 1460))
    #button.pack()
    #button = tk.Button(chartFrame, text='3 mths', command=lambda: create_popup(stock, 1733))
    #button.pack()
    #button = tk.Button(chartFrame, text='1 wk', command=lambda: create_popup(stock, 1818))
    #button.pack()

    button = tk.Button(chartFrame, text='5 yr', command=lambda: create_popup(stock, 252*5))
    button.pack()
    button = tk.Button(chartFrame, text='1 yr', command=lambda: create_popup(stock, 252))
    button.pack()
    button = tk.Button(chartFrame, text='3 mths', command=lambda: create_popup(stock, 252//4))
    button.pack()
    button = tk.Button(chartFrame, text='1 wk', command=lambda: create_popup(stock, 252//52))
    button.pack()
    df = df.iloc[-length:]

    # Find the columns and plot stock
    ax.plot(df['Date'], df['Closing'])
    for i in range(1, len(df)-length):
        ax.plot(df['Date'].iloc[i-1:i+1], df['Closing'].iloc[i-1:i+1], color=colors[i-1])

    prediction = predict(df, length)
    ax.plot(prediction[0], prediction[1], color='blue')
    ax.plot(prediction[2], prediction[3], color='orange')

    if not prediction[2].empty:
        ax.set_xlim([min(prediction[0]), max(prediction[2])])
    else:
        ax.set_xlim([min(prediction[0]), min(prediction[0])])


    canvas = FigureCanvasTkAgg(fig, master=chartFrame) 
    canvas.draw()
    canvas.get_tk_widget().pack()


# formatting window
root = tk.Tk()
root.geometry('800x600')
root.title("Stocks")
root.configure(bg = 'darkgray')
chartFrame = tk.Frame(root, width=650, height=400)
chartFrame.grid(row=0, column=0, padx=10, pady=5)
buttonFrame = tk.Frame(root, width=100, height=400)
buttonFrame.grid(row=0, column=1, padx=10, pady=5)


# button array to store the buttons for stocks
buttons = []

# error message for incorrect ticker
error = tk.Label(buttonFrame, text="Not a valid ticker, try again")
stockFound = tk.Label(buttonFrame, text="Stock updated/ added")

# verifies the ticker is an actual stock and if so then c
def verifyStock():
    ticker = entry.get()
    yf.Ticker(ticker).info
    try:
        yf.Ticker(ticker).info
        getStockData(ticker)
        for button in buttons:
            button.pack_forget()
        createButtons()
        error.pack_forget()
        stockFound.pack()
    except:
        error.pack_forget()
        error.pack()

# get new stocks
entry = tk.Entry(buttonFrame)
entry.pack()
button = tk.Button(buttonFrame, text='Get Stock', command=verifyStock)
button.pack()



# creates the buttons
def createButtons():
    files = os.listdir('./Data')

    stocks = []
    for i in range(len(files)):
        lenFileName = len(files[i])
        stocks.append(files[i][0:lenFileName-9])

    for stock in stocks:
        pctChange = round(pd.ExcelFile('./Data/' + stock + 'Data.xlsx').parse('Sheet')['Pct Change'].iloc[-1], 2)
        if pctChange > 0:
            button = tk.Button(buttonFrame, text=stock + ' ' + str(pctChange) + '%', command=lambda stock=stock: create_popup(stock, 0), fg = 'green')
        else:
            button = tk.Button(buttonFrame, text=stock + ' ' + str(pctChange) + '%', command=lambda stock=stock: create_popup(stock, 0), fg = 'red')
        button.pack()
        buttons.append(button)

createButtons()

# ----------------------------------------------------------------------------------------------------------------------
# prediction, technically uses polynomial but something is wrong, probably lacking features as its only using dates
def predict(df, length):

    df.set_index('Date', inplace=True)

    df = df.dropna()

    X = df.index.map(lambda date: date.toordinal()).values.reshape(-1, 1)
    y = df['Closing']

    poly = PolynomialFeatures(degree=1, include_bias=False)
    X_poly = poly.fit_transform(X)

    X_train, X_test, y_train, y_test = train_test_split(X_poly, y, test_size=0.2, random_state=99)

    X_test_dates = X_test[:, 0]

    model = LinearRegression()
    model.fit(X_train, y_train)

    r_sq = model.score(X_test, y_test)
    print('Coefficient of determination (R-squared):', r_sq)

    y_pred = model.predict(X_test)
    dates = [datetime.fromordinal(int(date)) for date in X_test_dates.flatten()]


    future_dates = pd.date_range(start=df.index.max(), periods=length/5)

    # Convert future dates to the same numerical format used for training
    future_dates_poly = np.array([])

    if len(future_dates) > 0:
        future_dates_num = future_dates.map(lambda date: date.toordinal()).values.reshape(-1, 1)
        future_dates_poly = poly.transform(future_dates_num)

    future_predictions = model.predict(future_dates_poly) if future_dates_poly.size else []

    return dates, y_pred, future_dates, future_predictions


root.mainloop()

# Add new file to add a predicting line and predict price in a year



